"""
Dashboard Financeiro Pessoal — v1 (com editor)
───────────────────────────────────────────────
Visualização + edição de transações com salvamento direto no Excel.

Requisitos:
    pip install streamlit plotly pandas openpyxl

Execução:
    streamlit run dashboard.py
"""

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from datetime import date
from pathlib import Path
import tempfile, shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Configuração ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Dashboard Financeiro",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    [data-testid="stSidebar"] { background-color: #141726; }
    .section-title {
        font-size: 17px; font-weight: 600; color: #7c83ff;
        margin: 20px 0 8px; border-bottom: 1px solid #2d3250; padding-bottom: 5px;
    }
    .save-box {
        background: #1e2130; border-radius: 10px; padding: 16px;
        border-left: 4px solid #4caf7d; margin: 12px 0;
    }
    .warn-box {
        background: #1e2130; border-radius: 10px; padding: 16px;
        border-left: 4px solid #f5a623; margin: 12px 0;
    }
    .empty-state { text-align:center; padding:50px 20px;
        border:2px dashed #2d3250; border-radius:16px; margin:30px auto; max-width:640px; }
    .empty-state h2 { color:#7c83ff; }
    .empty-state p  { color:#a0aab4; }
    .step-box { background:#1e2130; border-radius:10px; padding:16px 20px;
        margin:8px 0; border-left:4px solid #7c83ff; }
    .step-box b { color:#7c83ff; }
</style>
""", unsafe_allow_html=True)

COLORS = dict(primary="#7c83ff", success="#4caf7d", danger="#f05454",
              warning="#f5a623", secondary="#a0aab4", bg="#1e2130", grid="#2d3250")
PAL = ["#7c83ff","#4caf7d","#f05454","#f5a623","#56cfe1","#ff7096","#c77dff",
       "#06d6a0","#ff9f1c","#2ec4b6","#e76f51","#457b9d","#a8dadc","#ffd166","#ef476f"]
GROUP_LABELS = {"D.P.":"Despesas Pessoais","D.T.":"Transporte/Fixas",
                "D.F.":"Financeiras","PGT.":"Pagamentos","Vend":"Vendas/Receitas","Outr":"Outros"}
CURRENT_YEAR = date.today().year

# ─── Helpers para salvar no Excel ─────────────────────────────────────────────
def _side(): return Side(style="thin", color="C5C9D6")
def _bdr():  return Border(top=_side(), bottom=_side(), left=_side(), right=_side())
def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="1E2130"): return Font(name="Arial", size=10, bold=bold, color=color)
def _center(): return Alignment(horizontal="center", vertical="center")
def _left():   return Alignment(horizontal="left",   vertical="center")
def _right():  return Alignment(horizontal="right",  vertical="center")


# ─── Carregamento ─────────────────────────────────────────────────────────────
@st.cache_data
def load_data(filepath: str):
    xl = pd.ExcelFile(filepath)

    df = pd.read_excel(xl, "DB_DESPESAS", header=0)
    df.columns = df.columns.str.strip()
    df["Data Lançamento"] = pd.to_datetime(df["Data Lançamento"], errors="coerce")
    df["Data Base"]       = pd.to_datetime(df["Data Base"],       errors="coerce")
    df["Saída(R$)"]       = pd.to_numeric(df["Saída(R$)"],  errors="coerce").fillna(0)
    df["Entrada(R$)"]     = pd.to_numeric(df["Entrada(R$)"],errors="coerce").fillna(0)
    df["AnoMes"]          = df["Data Base"].dt.to_period("M").astype(str)
    gc = "GRUPO REAL" if "GRUPO REAL" in df.columns else "GRUPO"
    df["GRUPO LABEL"] = df[gc].map(GROUP_LABELS).fillna(df[gc])

    bdf = pd.read_excel(xl, "BD_BudgetPessoal", header=0)
    bdf.columns = bdf.columns.str.strip()
    bdf["Data Contábil"]    = pd.to_datetime(bdf["Data Contábil"], errors="coerce")
    bdf["Entrada Real"]     = pd.to_numeric(bdf["Entrada Real"],     errors="coerce").fillna(0)
    bdf["Entrada Esperada"] = pd.to_numeric(bdf["Entrada Esperada"], errors="coerce").fillna(0)

    def parse_period(val):
        try:
            parts = str(val).split("/")
            return pd.Period(year=int(parts[1]), month=int(parts[0]), freq="M").strftime("%Y-%m")
        except:
            return None
    bdf["AnoMes"] = bdf["Data Base"].apply(parse_period)

    regras = pd.read_excel(xl, "REGRAS_CATEGORIAS", header=0)
    regras.columns = regras.columns.str.strip()

    return df, bdf, regras, gc


# ─── Salvar edições no Excel ───────────────────────────────────────────────────
def save_edits_to_excel(filepath: str, edited_df: pd.DataFrame, deleted_ids: set):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["DB_DESPESAS"]

    id_to_row = {}
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row_idx, 1).value
        if val is not None:
            try:
                id_to_row[int(val)] = row_idx
            except (ValueError, TypeError):
                pass

    DB_COLS = ["ID","Data Lançamento","DESCRIÇÃO","Entrada(R$)","Saída(R$)",
               "CC","DESC. BASE","CATEGORIA","GRUPO","GRUPO REAL","STATUS","Data Base"]

    updated = 0
    for _, row in edited_df.iterrows():
        rid = int(row.get("ID", 0) or 0)
        if rid in deleted_ids or rid not in id_to_row:
            continue
        ri  = id_to_row[rid]
        alt = ri % 2 == 0
        bg  = "F0F2FF" if alt else "FFFFFF"

        for ci, col in enumerate(DB_COLS, 1):
            val  = row.get(col)
            safe = val if not (isinstance(val, float) and pd.isna(val)) else None
            cell = ws.cell(ri, ci, safe)
            cell.font = _font(); cell.fill = _fill(bg); cell.border = _bdr()
            if ci == 1:   cell.alignment = _center(); cell.font = _font(color="888888")
            elif ci == 2: cell.number_format = "DD/MM/YYYY"; cell.alignment = _center()
            elif ci in (4,5): cell.number_format = "#,##0.00"; cell.alignment = _right()
            elif ci in (9,10,11): cell.alignment = _center()
            elif ci == 12: cell.number_format = "DD/MM/YYYY"; cell.alignment = _center()
            if ci == 11:
                sv = str(val).upper() if val else ""
                if sv == "PAGO":
                    cell.font = Font(name="Arial", size=10, bold=True, color="4CAF7D")
                elif sv == "PENDENTE":
                    cell.font = Font(name="Arial", size=10, bold=True, color="F05454")
        updated += 1

    deleted = 0
    if deleted_ids:
        rows_to_delete = sorted(
            [id_to_row[rid] for rid in deleted_ids if rid in id_to_row],
            reverse=True
        )
        for ri in rows_to_delete:
            ws.delete_rows(ri)
            deleted += 1

    wb.save(filepath)
    return updated, deleted


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 📂 Arquivo de Dados")
    uploaded = st.file_uploader("Upload do Excel (.xlsx)", type=["xlsx"])

    filepath = None; df = None; bdf = None; regras = None
    gc = "GRUPO REAL"; load_error = None
    default_file = Path(__file__).parent / "controle_financeiro.xlsx"

    if uploaded:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        shutil.copyfileobj(uploaded, tmp); tmp.flush()
        filepath = tmp.name
    elif default_file.exists():
        filepath = str(default_file)

    if filepath:
        try:
            df, bdf, regras, gc = load_data(filepath)
        except Exception as e:
            load_error = str(e)

    if df is not None:
        st.markdown("---")
        st.markdown("## 🔍 Filtros")
        all_months = sorted(df["AnoMes"].dropna().unique())

        def pick(months, target, fallback):
            return months.index(target) if target in months else fallback

        i_s = pick(all_months, f"{CURRENT_YEAR}-01", 0)
        i_e = pick(all_months, f"{CURRENT_YEAR}-12", len(all_months) - 1)

        c1, c2 = st.columns(2)
        start_m = c1.selectbox("De",  all_months, index=i_s)
        end_m   = c2.selectbox("Até", all_months, index=i_e)
        valid_range = [m for m in all_months if start_m <= m <= end_m]

        all_groups  = sorted(df[gc].dropna().unique())
        sel_groups  = st.multiselect("Grupos",     all_groups,  default=all_groups)
        all_cats    = sorted(df["CATEGORIA"].dropna().unique())
        sel_cats    = st.multiselect("Categorias", all_cats,    default=all_cats)
        status_opts = sorted(df["STATUS"].dropna().unique())
        sel_status  = st.multiselect("Status",     status_opts, default=status_opts)
        st.markdown("---")
        st.caption(f"📅 Filtro padrão: **{CURRENT_YEAR}**")


# ══════════════════════════════════════════════════════════════════════════════
# ESTADO ZERO
# ══════════════════════════════════════════════════════════════════════════════
if df is None:
    st.title("💰 Dashboard Financeiro Pessoal")
    if load_error:
        st.error(f"❌ {load_error}")
    else:
        st.markdown("""<div class="empty-state">
            <h2>Bem-vindo! 👋</h2>
            <p>Faça upload do <b>controle_financeiro.xlsx</b> na barra lateral,<br>
            ou coloque-o na mesma pasta deste script — ele carrega automaticamente.</p>
        </div>""", unsafe_allow_html=True)

    st.markdown('<div class="section-title">⚙️ Fluxo de Uso</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
        <div class="step-box"><b>① Cole os extratos</b><br>
        Abas <code>RAW_C6</code>, <code>RAW_BRADESCO</code>, <code>RAW_NUBANK</code>, <code>RAW_INTER</code></div>
        <div class="step-box"><b>② Rode o processador</b><br>
        <code>python processar_extratos.py</code><br>
        Detecta novas linhas, deduplica e insere em DB_DESPESAS.</div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
        <div class="step-box"><b>③ Abra o dashboard</b><br>
        <code>streamlit run dashboard.py</code><br>
        Revise e corrija via aba <b>✏️ Editor</b>.</div>
        <div class="step-box"><b>④ Salve de volta</b><br>
        Clique <b>Salvar no Excel</b> — edições voltam para o <code>.xlsx</code>
        e as regras novas valem no próximo mês.</div>""", unsafe_allow_html=True)
    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
# FILTRO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
mask = (df["AnoMes"].isin(valid_range) & df[gc].isin(sel_groups) &
        df["CATEGORIA"].isin(sel_cats)  & df["STATUS"].isin(sel_status))
fdf       = df[mask].copy()
bdf_range = bdf[bdf["AnoMes"].isin(valid_range)].copy()

st.title("💰 Dashboard Financeiro Pessoal")
st.caption(f"Período: **{start_m}** → **{end_m}** · {len(valid_range)} meses · {len(fdf):,} transações")

sem_cat = df[df["AnoMes"].isin(valid_range) &
             (df["CATEGORIA"].isna() | (df["CATEGORIA"].str.strip() == ""))]
if not sem_cat.empty:
    st.markdown(f"""<div class="warn-box">
    ⚠️ <b>{len(sem_cat)} transação(ões) sem categoria</b> no período.
    Use a aba <b>✏️ Editor</b> para classificá-las.
    </div>""", unsafe_allow_html=True)

tab_dash, tab_budget, tab_editor = st.tabs(["📊 Dashboard", "🎯 Budget", "✏️ Editor"])


# ══════════════════════════════════════════════════════════════════════════════
# TAB DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
with tab_dash:
    st.markdown('<div class="section-title">📊 Resumo do Período</div>', unsafe_allow_html=True)
    total_saida   = fdf["Saída(R$)"].abs().sum()
    total_entrada = fdf["Entrada(R$)"].sum()
    saldo         = total_entrada - total_saida
    pago          = fdf[fdf["STATUS"]=="PAGO"]["Saída(R$)"].abs().sum()
    pendente      = fdf[fdf["STATUS"]=="PENDENTE"]["Saída(R$)"].abs().sum()
    media_mensal  = (fdf.groupby("AnoMes")["Saída(R$)"]
                     .apply(lambda x: x.abs().sum()).mean()) if valid_range else 0

    k = st.columns(6)
    k[0].metric("💸 Total Saídas",   f"R$ {total_saida:,.0f}")
    k[1].metric("💵 Total Entradas", f"R$ {total_entrada:,.0f}")
    k[2].metric("📈 Saldo Líquido",  f"R$ {saldo:,.0f}",    delta=f"R$ {saldo:,.0f}")
    k[3].metric("✅ Pago",           f"R$ {pago:,.0f}")
    k[4].metric("⏳ Pendente",       f"R$ {pendente:,.0f}", delta_color="inverse")
    k[5].metric("📅 Média Mensal",   f"R$ {media_mensal:,.0f}")

    st.markdown('<div class="section-title">📅 Evolução Mensal por Grupo</div>', unsafe_allow_html=True)
    mg = (fdf.groupby(["AnoMes","GRUPO LABEL"])["Saída(R$)"]
          .apply(lambda x: x.abs().sum()).reset_index().sort_values("AnoMes"))
    if not mg.empty:
        fig = px.bar(mg, x="AnoMes", y="Saída(R$)", color="GRUPO LABEL",
                     color_discrete_sequence=PAL, template="plotly_dark",
                     labels={"AnoMes":"Mês","Saída(R$)":"R$","GRUPO LABEL":"Grupo"})
        fig.update_layout(plot_bgcolor=COLORS["bg"], paper_bgcolor=COLORS["bg"],
                          barmode="stack", legend=dict(orientation="h",y=-0.22),
                          xaxis=dict(tickangle=-45), margin=dict(t=10,b=70))
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Sem dados no período.")

    st.markdown('<div class="section-title">🏷️ Categorias e Grupos</div>', unsafe_allow_html=True)
    ca, cb = st.columns([1.3, 1])
    with ca:
        cs = (fdf.groupby("CATEGORIA")["Saída(R$)"]
              .apply(lambda x: x.abs().sum()).sort_values().reset_index())
        if not cs.empty:
            fig2 = px.bar(cs, x="Saída(R$)", y="CATEGORIA", orientation="h",
                          color="Saída(R$)", color_continuous_scale="Blues",
                          labels={"Saída(R$)":"R$","CATEGORIA":""}, template="plotly_dark")
            fig2.update_layout(plot_bgcolor=COLORS["bg"], paper_bgcolor=COLORS["bg"],
                               coloraxis_showscale=False, height=460,
                               margin=dict(t=10,b=10,l=10,r=10))
            st.plotly_chart(fig2, use_container_width=True)
    with cb:
        gs = (fdf.groupby("GRUPO LABEL")["Saída(R$)"]
              .apply(lambda x: x.abs().sum()).reset_index())
        if not gs.empty:
            fig3 = px.pie(gs, values="Saída(R$)", names="GRUPO LABEL",
                          color_discrete_sequence=PAL, hole=0.45, template="plotly_dark")
            fig3.update_traces(textposition="outside", textinfo="percent+label")
            fig3.update_layout(plot_bgcolor=COLORS["bg"], paper_bgcolor=COLORS["bg"],
                               showlegend=False, height=460,
                               margin=dict(t=10,b=10,l=10,r=10))
            st.plotly_chart(fig3, use_container_width=True)

    st.markdown('<div class="section-title">🌡️ Heatmap: Categoria × Mês</div>', unsafe_allow_html=True)
    if not fdf.empty:
        heat = (fdf.groupby(["AnoMes","CATEGORIA"])["Saída(R$)"]
                .apply(lambda x: x.abs().sum()).unstack(fill_value=0).sort_index())
        heat = heat[sorted(heat.columns)]
        fig4 = go.Figure(go.Heatmap(
            z=heat.values, x=heat.columns.tolist(), y=heat.index.tolist(),
            colorscale="Blues",
            text=[[f"R$ {v:,.0f}" for v in r] for r in heat.values],
            texttemplate="%{text}", textfont={"size":9}, hoverongaps=False))
        fig4.update_layout(plot_bgcolor=COLORS["bg"], paper_bgcolor=COLORS["bg"],
                           height=420, xaxis=dict(tickangle=-45),
                           font=dict(color="#e8eaf0"), margin=dict(t=10,b=10))
        st.plotly_chart(fig4, use_container_width=True)

    st.markdown('<div class="section-title">📉 Fluxo Mensal: Despesas × Receitas</div>', unsafe_allow_html=True)
    if not fdf.empty:
        dm = (fdf.groupby("AnoMes")["Saída(R$)"]
              .apply(lambda x: x.abs().sum()).reset_index()
              .rename(columns={"Saída(R$)":"Despesas"}))
        em = (fdf.groupby("AnoMes")["Entrada(R$)"].sum().reset_index()
              .rename(columns={"Entrada(R$)":"Receitas"}))
        fx = dm.merge(em, on="AnoMes", how="outer").sort_values("AnoMes").fillna(0)
        fx["Saldo"] = fx["Receitas"] - fx["Despesas"]
        fig5 = go.Figure()
        fig5.add_trace(go.Scatter(x=fx["AnoMes"], y=fx["Despesas"], name="Despesas",
                                  mode="lines+markers",
                                  line=dict(color=COLORS["danger"], width=2.5)))
        fig5.add_trace(go.Scatter(x=fx["AnoMes"], y=fx["Receitas"], name="Receitas",
                                  mode="lines+markers",
                                  line=dict(color=COLORS["success"], width=2.5)))
        fig5.add_trace(go.Bar(x=fx["AnoMes"], y=fx["Saldo"], name="Saldo", opacity=0.4,
                              marker_color=[COLORS["success"] if v >= 0 else COLORS["danger"]
                                            for v in fx["Saldo"]]))
        fig5.update_layout(plot_bgcolor=COLORS["bg"], paper_bgcolor=COLORS["bg"],
                           legend=dict(orientation="h",y=-0.22),
                           xaxis=dict(tickangle=-45), yaxis_title="R$",
                           margin=dict(t=10,b=70), font=dict(color="#e8eaf0"),
                           template="plotly_dark")
        st.plotly_chart(fig5, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB BUDGET
# ══════════════════════════════════════════════════════════════════════════════
with tab_budget:
    total_real     = bdf_range["Entrada Real"].sum()
    total_esperado = bdf_range["Entrada Esperada"].sum()
    budget_gap     = total_real - total_esperado

    st.markdown('<div class="section-title">🎯 Entradas — Real vs Esperado</div>',
                unsafe_allow_html=True)
    bk = st.columns(3)
    bk[0].metric("💰 Entrada Real",         f"R$ {total_real:,.0f}")
    bk[1].metric("🎯 Entrada Esperada",     f"R$ {total_esperado:,.0f}")
    bk[2].metric("📊 Gap Real vs Esperado", f"R$ {budget_gap:,.0f}",
                 delta=f"R$ {budget_gap:,.0f}")

    bm = (bdf_range.groupby("AnoMes")[["Entrada Real","Entrada Esperada"]]
          .sum().reset_index().sort_values("AnoMes"))

    if not bm.empty:
        fig6 = go.Figure()
        fig6.add_trace(go.Bar(x=bm["AnoMes"], y=bm["Entrada Esperada"], name="Esperado",
                              marker_color=COLORS["secondary"], opacity=0.6))
        fig6.add_trace(go.Bar(x=bm["AnoMes"], y=bm["Entrada Real"],     name="Real",
                              marker_color=COLORS["success"]))
        fig6.update_layout(barmode="group", plot_bgcolor=COLORS["bg"],
                           paper_bgcolor=COLORS["bg"],
                           legend=dict(orientation="h",y=-0.22),
                           xaxis=dict(tickangle=-45), yaxis_title="R$",
                           margin=dict(t=10,b=70), font=dict(color="#e8eaf0"),
                           template="plotly_dark")
        st.plotly_chart(fig6, use_container_width=True)

        cc, cd = st.columns([1, 1])
        with cc:
            bt = (bdf_range.groupby("Título")[["Entrada Real","Entrada Esperada"]]
                  .sum().sort_values("Entrada Real", ascending=False).reset_index())
            fig7 = go.Figure()
            fig7.add_trace(go.Bar(x=bt["Título"], y=bt["Entrada Esperada"], name="Esperado",
                                  marker_color=COLORS["secondary"], opacity=0.6))
            fig7.add_trace(go.Bar(x=bt["Título"], y=bt["Entrada Real"], name="Real",
                                  marker_color=COLORS["success"]))
            fig7.update_layout(barmode="group", title="Por Título",
                               plot_bgcolor=COLORS["bg"], paper_bgcolor=COLORS["bg"],
                               xaxis=dict(tickangle=-35),
                               legend=dict(orientation="h",y=-0.3),
                               margin=dict(t=40,b=80), height=380,
                               font=dict(color="#e8eaf0"), template="plotly_dark")
            st.plotly_chart(fig7, use_container_width=True)
        with cd:
            pct = (total_real / total_esperado * 100) if total_esperado > 0 else 0
            fig8 = go.Figure(go.Indicator(
                mode="gauge+number+delta", value=pct,
                number={"suffix":"%","font":{"color":"#e8eaf0"}},
                delta={"reference":100,"suffix":"%"},
                title={"text":"% do Budget Realizado",
                       "font":{"color":"#e8eaf0","size":14}},
                gauge={
                    "axis":{"range":[0,120],"tickcolor":"#a0aab4"},
                    "bar":{"color":COLORS["primary"]},
                    "steps":[{"range":[0,70],   "color":"#f05454"},
                              {"range":[70,90],  "color":"#f5a623"},
                              {"range":[90,105], "color":"#4caf7d"},
                              {"range":[105,120],"color":"#7c83ff"}],
                    "threshold":{"line":{"color":"white","width":3},
                                 "thickness":0.75,"value":100},
                    "bgcolor":COLORS["bg"],
                },))
            fig8.update_layout(paper_bgcolor=COLORS["bg"],
                               font=dict(color="#e8eaf0"),
                               margin=dict(t=40,b=20,l=20,r=20), height=330)
            st.plotly_chart(fig8, use_container_width=True)

        st.markdown('<div class="section-title">📋 Detalhamento Budget</div>',
                    unsafe_allow_html=True)
        bshow = bdf_range.copy()
        bshow["Data Contábil"]    = bshow["Data Contábil"].dt.strftime("%d/%m/%Y")
        bshow["Entrada Real"]     = bshow["Entrada Real"].apply(lambda x: f"R$ {x:,.2f}")
        bshow["Entrada Esperada"] = bshow["Entrada Esperada"].apply(lambda x: f"R$ {x:,.2f}")
        st.dataframe(
            bshow[["Data Contábil","AnoMes","Título","Entrada Real","Entrada Esperada"]],
            use_container_width=True, height=320)
    else:
        st.info("Sem dados de budget no período.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB EDITOR
# ══════════════════════════════════════════════════════════════════════════════
with tab_editor:
    st.markdown('<div class="section-title">✏️ Editor de Transações</div>',
                unsafe_allow_html=True)

    ef1, ef2, ef3 = st.columns(3)
    filtro_sem_cat  = ef1.checkbox("🔶 Só sem categoria", value=False)
    filtro_pendente = ef2.checkbox("⏳ Só PENDENTE",      value=False)
    busca_desc      = ef3.text_input("🔍 Buscar descrição", placeholder="ex: ABRAKABECA")

    ed_df = df[df["AnoMes"].isin(valid_range)].copy()
    if filtro_sem_cat:
        ed_df = ed_df[ed_df["CATEGORIA"].isna() | (ed_df["CATEGORIA"].str.strip() == "")]
    if filtro_pendente:
        ed_df = ed_df[ed_df["STATUS"].str.upper() == "PENDENTE"]
    if busca_desc.strip():
        ed_df = ed_df[ed_df["DESCRIÇÃO"].str.contains(busca_desc.strip(),
                                                        case=False, na=False)]

    st.caption(f"{len(ed_df)} transação(ões) · Edite na tabela e clique **Salvar no Excel**.")

    all_cats_ed = sorted(df["CATEGORIA"].dropna().unique().tolist())
    all_grp_ed  = ["D.P.","D.T.","D.F.","PGT.","Vend","Outr"]
    all_sts_ed  = ["PAGO","PENDENTE"]

    edit_cols = ["ID","Data Lançamento","DESCRIÇÃO","Saída(R$)","Entrada(R$)",
                 "CC","DESC. BASE","CATEGORIA","GRUPO","GRUPO REAL","STATUS","Data Base","AnoMes"]
    display_df = ed_df[[c for c in edit_cols if c in ed_df.columns]].copy()

    edited = st.data_editor(
        display_df,
        use_container_width=True,
        height=480,
        num_rows="dynamic",
        column_config={
            "ID":              st.column_config.NumberColumn("ID", disabled=True, width="small"),
            "Data Lançamento": st.column_config.DateColumn("Data Lanç.", format="DD/MM/YYYY"),
            "DESCRIÇÃO":       st.column_config.TextColumn("Descrição", width="large"),
            "Saída(R$)":       st.column_config.NumberColumn("Saída(R$)",   format="R$ %.2f"),
            "Entrada(R$)":     st.column_config.NumberColumn("Entrada(R$)", format="R$ %.2f"),
            "CC":              st.column_config.TextColumn("CC", width="small"),
            "DESC. BASE":      st.column_config.TextColumn("DESC. BASE", width="medium"),
            "CATEGORIA":       st.column_config.SelectboxColumn("Categoria",
                                   options=all_cats_ed, width="medium"),
            "GRUPO":           st.column_config.SelectboxColumn("Grupo",
                                   options=all_grp_ed, width="small"),
            "GRUPO REAL":      st.column_config.SelectboxColumn("Grupo Real",
                                   options=all_grp_ed, width="small"),
            "STATUS":          st.column_config.SelectboxColumn("Status",
                                   options=all_sts_ed, width="small"),
            "Data Base":       st.column_config.DateColumn("Data Base",  format="DD/MM/YYYY"),
            "AnoMes":          st.column_config.TextColumn("AnoMes", disabled=True, width="small"),
        },
        key="editor_table",
    )

    if "deleted_ids" not in st.session_state:
        st.session_state.deleted_ids = set()

    original_ids = set(display_df["ID"].dropna().astype(int).tolist())
    edited_ids   = (set(edited["ID"].dropna().astype(int).tolist())
                    if "ID" in edited.columns else original_ids)
    newly_deleted = original_ids - edited_ids
    if newly_deleted:
        st.session_state.deleted_ids.update(newly_deleted)

    if st.session_state.deleted_ids:
        st.warning(f"🗑️ {len(st.session_state.deleted_ids)} linha(s) marcada(s) para exclusão."
                   " Clique **Salvar** para confirmar.")

    st.markdown("---")
    cs1, cs2, cs3 = st.columns([2, 1, 1])
    with cs1:
        st.markdown('<div class="save-box">💾 As edições são gravadas diretamente no <code>.xlsx</code>.'
                    ' Mantenha um backup antes de grandes alterações.</div>',
                    unsafe_allow_html=True)
    with cs2:
        if st.button("💾 Salvar no Excel", type="primary", use_container_width=True):
            if not filepath or not Path(filepath).exists():
                st.error("Arquivo não encontrado para salvar.")
            else:
                try:
                    upd, dlt = save_edits_to_excel(
                        filepath, edited, st.session_state.deleted_ids)
                    st.session_state.deleted_ids.clear()
                    st.cache_data.clear()
                    st.success(f"✅ {upd} atualizada(s), {dlt} excluída(s). Recarregando…")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Erro ao salvar: {e}")
    with cs3:
        if st.button("↩️ Descartar", use_container_width=True):
            st.session_state.deleted_ids.clear()
            st.rerun()

    with st.expander("📋 Tabela de Regras (REGRAS_CATEGORIAS)", expanded=False):
        st.caption("Edite diretamente no Excel. "
                   "O script processar_extratos.py usa estas regras na próxima importação.")
        if regras is not None and not regras.empty:
            st.dataframe(regras, use_container_width=True, height=300)
        else:
            st.info("Tabela de regras não encontrada.")

# Footer
st.markdown("---")
st.caption(f"Dashboard Financeiro · Ano padrão: {CURRENT_YEAR} · controle_financeiro.xlsx")
