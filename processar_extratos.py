"""
processar_extratos.py
─────────────────────
Lê as abas RAW_ do controle_financeiro.xlsx, normaliza cada banco,
aplica as REGRAS_CATEGORIAS e atualiza DB_DESPESAS com as novas linhas.

Uso:
    python processar_extratos.py
    python processar_extratos.py --arquivo meu_controle.xlsx
    python processar_extratos.py --banco C6          # processa só o C6
    python processar_extratos.py --banco NUBANK,INTER

Requisitos:
    pip install pandas openpyxl
"""

import argparse
import sys
from pathlib import Path
from datetime import date
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Defaults ─────────────────────────────────────────────────────────────────
DEFAULT_FILE = Path(__file__).parent / "controle_financeiro.xlsx"
BANCOS_DISPONIVEIS = ["C6", "BRADESCO", "NUBANK", "INTER"]

# ─── Estilo helpers ────────────────────────────────────────────────────────────
def _side(): return Side(style="thin", color="C5C9D6")
def _border(): return Border(top=_side(), bottom=_side(), left=_side(), right=_side())
def _fill(c): return PatternFill("solid", fgColor=c)
def _bfont(bold=False, color="1E2130"): return Font(name="Arial", size=10, bold=bold, color=color)
def _center(): return Alignment(horizontal="center", vertical="center")
def _left(): return Alignment(horizontal="left", vertical="center")
def _right(): return Alignment(horizontal="right", vertical="center")

def _style_data_row(ws, row_idx, ncols, alt=False):
    bg = "F0F2FF" if alt else "FFFFFF"
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = _bfont()
        cell.fill = _fill(bg)
        cell.alignment = _left()
        cell.border = _border()
    ws.row_dimensions[row_idx].height = 19


# ══════════════════════════════════════════════════════════════════════════════
# NORMALIZADORES — cada banco tem seu próprio parser
# Todos retornam DataFrame com colunas padrão:
#   Data Lançamento | DESCRIÇÃO | Entrada(R$) | Saída(R$) | CC
# ══════════════════════════════════════════════════════════════════════════════

def _clean_valor(series):
    """
    Converte série para numérico tratando formato BR (1.234,56) e US (1234.56).
    Também trata células já numéricas (float/int) vindas direto do Excel.
    """
    # Se já for numérico (Excel leu como número), devolve direto
    if pd.api.types.is_numeric_dtype(series):
        return series.fillna(0)

    s = series.astype(str).str.strip()
    s = s.str.replace(r"\s", "", regex=True)   # espaços internos
    s = s.str.replace("R$", "", regex=False)
    s = s.str.replace("\xa0", "", regex=False)  # non-breaking space

    # Detecta formato BR: tem ponto de milhar E vírgula decimal  → 1.234,56
    # ou só vírgula decimal → 1234,56
    # Formato US: só ponto decimal → 1234.56
    has_dot_and_comma = s.str.contains(r"\d\.\d{3},", regex=True)
    has_only_comma    = s.str.contains(r",") & ~s.str.contains(r"\.")

    # Formato BR: remove ponto, troca vírgula por ponto
    br_mask = has_dot_and_comma | has_only_comma
    s_br = s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
    s_us = s.str.replace(",", "", regex=False)  # remove separador de milhar US

    # Aplica conversão: BR onde br_mask, senão tenta US
    s_final = s.copy()
    s_final[br_mask]  = s_br[br_mask]
    s_final[~br_mask] = s_us[~br_mask]
    return pd.to_numeric(s_final, errors="coerce").fillna(0)


def _find_col(cols, *keywords):
    """
    Procura a primeira coluna cujo nome (string) contenha qualquer keyword.
    Ignora colunas cujo nome seja NaN/float.
    Retorna None se não encontrar.
    """
    str_cols = [c for c in cols if isinstance(c, str)]
    for kw in keywords:
        match = next((c for c in str_cols if kw in c.lower()), None)
        if match:
            return match
    return None


def _get_col(df, *keywords, default=None):
    """Retorna a Series da coluna encontrada, ou uma Series de zeros/vazia."""
    col = _find_col(df.columns, *keywords)
    if col:
        return df[col]
    if default is not None:
        return pd.Series([default] * len(df), index=df.index)
    return pd.Series([""] * len(df), index=df.index)


def normalizar_c6(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    C6 exporta: Data Lançamento | Data Contábil | Título | Descrição |
                Entrada(R$) | Saída(R$) | Saldo do Dia(R$)
    Junta Título + Descrição como DESCRIÇÃO.
    """
    df = df_raw.copy()
    df.columns = df.columns.str.strip()

    data_col = _find_col(df.columns, "lança", "data") or df.columns[0]
    df["Data Lançamento"] = pd.to_datetime(df[data_col], dayfirst=True, errors="coerce")

    titulo = _get_col(df, "título", "titulo").fillna("").astype(str).str.strip()
    descr  = _get_col(df, "descriç", "descric").fillna("").astype(str).str.strip()
    df["DESCRIÇÃO"] = (titulo + " | " + descr).str.strip(" |")
    df["DESCRIÇÃO"] = df["DESCRIÇÃO"].str.replace(r"^\s*\|\s*|\s*\|\s*$", "", regex=True)

    df["Entrada(R$)"] = _clean_valor(_get_col(df, "entrada", default=0))
    df["Saída(R$)"]   = _clean_valor(_get_col(df, "saída", "saida", default=0)).abs() * -1
    df["CC"]          = "C6"
    return df[["Data Lançamento", "DESCRIÇÃO", "Entrada(R$)", "Saída(R$)", "CC"]].dropna(subset=["Data Lançamento"])


def normalizar_bradesco(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Bradesco exporta: Data Contábil | Título | Descrição |
                      Entrada(R$) | Saída(R$) | Saldo do Dia(R$)
    """
    df = df_raw.copy()
    df.columns = df.columns.str.strip()

    data_col = _find_col(df.columns, "data") or df.columns[0]
    df["Data Lançamento"] = pd.to_datetime(df[data_col], dayfirst=True, errors="coerce")

    titulo = _get_col(df, "título", "titulo").fillna("").astype(str).str.strip()
    descr  = _get_col(df, "descriç", "descric").fillna("").astype(str).str.strip()
    df["DESCRIÇÃO"] = (titulo + " | " + descr).str.strip(" |")
    df["DESCRIÇÃO"] = df["DESCRIÇÃO"].str.replace(r"^\s*\|\s*|\s*\|\s*$", "", regex=True)

    df["Entrada(R$)"] = _clean_valor(_get_col(df, "entrada", default=0))
    df["Saída(R$)"]   = _clean_valor(_get_col(df, "saída", "saida", default=0)).abs() * -1
    df["CC"]          = "BRADESCO"
    return df[["Data Lançamento", "DESCRIÇÃO", "Entrada(R$)", "Saída(R$)", "CC"]].dropna(subset=["Data Lançamento"])


def normalizar_nubank(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Nubank CSV exporta: Data | Descrição | Valor
    Valor positivo = entrada, negativo = saída.
    Também suporta as colunas Entrada(R$) / Saída(R$) da RAW_NUBANK.
    """
    df = df_raw.copy()
    df.columns = df.columns.str.strip()

    data_col = _find_col(df.columns, "data") or df.columns[0]
    df["Data Lançamento"] = pd.to_datetime(df[data_col], dayfirst=True, errors="coerce")

    desc_col = _find_col(df.columns, "descriç", "descric") or df.columns[1]
    df["DESCRIÇÃO"] = df[desc_col].fillna("").astype(str).str.strip()

    valor_col = _find_col(df.columns, "valor")
    if valor_col:
        valor = _clean_valor(df[valor_col])
        df["Entrada(R$)"] = valor.clip(lower=0)
        df["Saída(R$)"]   = valor.clip(upper=0)
    else:
        df["Entrada(R$)"] = _clean_valor(_get_col(df, "entrada", default=0))
        df["Saída(R$)"]   = _clean_valor(_get_col(df, "saída", "saida", default=0)).abs() * -1

    df["CC"] = "NUBANK"
    return df[["Data Lançamento", "DESCRIÇÃO", "Entrada(R$)", "Saída(R$)", "CC"]].dropna(subset=["Data Lançamento"])


def normalizar_inter(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Inter exporta: Data Lançamento | Histórico | Descrição | Valor | Saldo
    Valor positivo = entrada, negativo = saída.
    Junta Histórico + Descrição como DESCRIÇÃO.
    """
    df = df_raw.copy()
    df.columns = df.columns.str.strip()

    data_col = _find_col(df.columns, "data") or df.columns[0]
    df["Data Lançamento"] = pd.to_datetime(df[data_col], dayfirst=True, errors="coerce")

    # Histórico é o tipo da transação; Descrição é o detalhe — junta os dois
    historico = _get_col(df, "históric", "histor").fillna("").astype(str).str.strip()
    descricao = _get_col(df, "descriç", "descric").fillna("").astype(str).str.strip()
    df["DESCRIÇÃO"] = (historico + " | " + descricao).str.strip(" |")
    df["DESCRIÇÃO"] = df["DESCRIÇÃO"].str.replace(r"^\s*\|\s*|\s*\|\s*$", "", regex=True)

    # Inter usa coluna única "Valor": positivo = crédito, negativo = débito
    valor_col = _find_col(df.columns, "valor")
    if valor_col:
        valor = _clean_valor(df[valor_col])
        df["Entrada(R$)"] = valor.clip(lower=0)
        df["Saída(R$)"]   = valor.clip(upper=0)
    else:
        df["Entrada(R$)"] = _clean_valor(_get_col(df, "entrada", default=0))
        df["Saída(R$)"]   = _clean_valor(_get_col(df, "saída", "saida", default=0)).abs() * -1

    df["CC"] = "INTER"
    return df[["Data Lançamento", "DESCRIÇÃO", "Entrada(R$)", "Saída(R$)", "CC"]].dropna(subset=["Data Lançamento"])


NORMALIZADORES = {
    "C6":       normalizar_c6,
    "BRADESCO": normalizar_bradesco,
    "NUBANK":   normalizar_nubank,
    "INTER":    normalizar_inter,
}


# ══════════════════════════════════════════════════════════════════════════════
# APLICAR REGRAS
# ══════════════════════════════════════════════════════════════════════════════

def aplicar_regras(df: pd.DataFrame, regras: pd.DataFrame) -> pd.DataFrame:
    """
    Junta o DataFrame normalizado com a tabela de regras por DESC. BASE.
    Também infere Data Base (1º dia do mês da Data Lançamento).
    Transações sem DESC. BASE ficam com CATEGORIA vazia para revisão manual.
    """
    df = df.copy()
    df["DESC. BASE"]  = ""
    df["CATEGORIA"]   = ""
    df["GRUPO"]       = ""
    df["GRUPO REAL"]  = ""
    df["STATUS"]      = "PAGO"
    df["Data Base"]   = df["Data Lançamento"].apply(
        lambda d: d.replace(day=1) if pd.notna(d) else pd.NaT
    )
    # Regras nunca são aplicadas automaticamente à DESC. BASE —
    # a DESC. BASE é preenchida pelo usuário; a CATEGORIA é derivada dela.
    # Aqui apenas pré-carregamos o mapeamento para uso futuro no dashboard.
    return df


# ══════════════════════════════════════════════════════════════════════════════
# DEDUPLICAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

def deduplicate(df_existing: pd.DataFrame, df_new: pd.DataFrame) -> pd.DataFrame:
    """
    Remove de df_new qualquer linha que já exista em df_existing,
    comparando Data Lançamento + DESCRIÇÃO + Saída(R$) + CC.
    """
    key_cols = ["Data Lançamento", "DESCRIÇÃO", "Saída(R$)", "CC"]

    def make_key(df):
        return (
            df["Data Lançamento"].dt.strftime("%Y-%m-%d").fillna("") + "|" +
            df["DESCRIÇÃO"].fillna("").str.strip().str.upper() + "|" +
            df["Saída(R$)"].fillna(0).astype(str) + "|" +
            df["CC"].fillna("")
        )

    existing_keys = set(make_key(df_existing))
    new_keys = make_key(df_new)
    mask = ~new_keys.isin(existing_keys)
    n_dupl = (~mask).sum()
    if n_dupl:
        print(f"  ⚠️  {n_dupl} linha(s) duplicada(s) ignorada(s).")
    return df_new[mask].copy()


# ══════════════════════════════════════════════════════════════════════════════
# LEITURA DA RAW SHEET
# ══════════════════════════════════════════════════════════════════════════════

def ler_raw_sheet(xl: pd.ExcelFile, banco: str) -> pd.DataFrame:
    """
    Lê a aba RAW_{banco}.
    Linha 1 = instrução (ignorada), Linha 2 = cabeçalho, Linha 3+ = dados.
    Retorna DataFrame vazio se não houver dados.
    """
    sheet = f"RAW_{banco}"
    if sheet not in xl.sheet_names:
        raise ValueError(f"Aba '{sheet}' não encontrada.")

    df_raw = pd.read_excel(xl, sheet_name=sheet, header=None)
    df_raw = df_raw.dropna(how="all")

    if len(df_raw) < 3:
        return pd.DataFrame()

    header_row = df_raw.iloc[1].tolist()
    data_rows  = df_raw.iloc[2:].reset_index(drop=True)
    data_rows.columns = [str(h).strip() if isinstance(h, str) else h for h in header_row]
    data_rows = data_rows.loc[:, data_rows.columns.notna()]
    data_rows = data_rows.loc[:, [c for c in data_rows.columns if str(c).strip() not in ("nan", "None", "")]]
    data_rows = data_rows.dropna(how="all")
    return data_rows


# ══════════════════════════════════════════════════════════════════════════════
# ESCRITA NO DB_DESPESAS
# ══════════════════════════════════════════════════════════════════════════════

DB_COLS = [
    "ID", "Data Lançamento", "DESCRIÇÃO", "Entrada(R$)", "Saída(R$)",
    "CC", "DESC. BASE", "CATEGORIA", "GRUPO", "GRUPO REAL", "STATUS", "Data Base"
]

def escrever_db_despesas(wb: openpyxl.Workbook, df_new: pd.DataFrame, max_existing_id: int):
    """Acrescenta linhas novas ao final da aba DB_DESPESAS."""
    ws = wb["DB_DESPESAS"]

    # Encontra próxima linha vazia
    next_row = ws.max_row + 1
    # Garante que não há linhas fantasma
    while next_row > 2 and ws.cell(next_row - 1, 1).value is None:
        next_row -= 1

    next_id = max_existing_id + 1

    for i, (_, row) in enumerate(df_new.iterrows()):
        ri = next_row + i
        alt = ri % 2 == 0
        _style_data_row(ws, ri, len(DB_COLS), alt=alt)

        vals = [
            next_id + i,
            row.get("Data Lançamento"),
            str(row.get("DESCRIÇÃO", "")).strip(),
            float(row.get("Entrada(R$)", 0) or 0),
            float(row.get("Saída(R$)", 0) or 0),
            str(row.get("CC", "")),
            str(row.get("DESC. BASE", "")),
            str(row.get("CATEGORIA", "")),
            str(row.get("GRUPO", "")),
            str(row.get("GRUPO REAL", "")),
            str(row.get("STATUS", "PAGO")),
            row.get("Data Base"),
        ]

        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v if not (isinstance(v, float) and pd.isna(v)) else None)
            if ci == 1:
                cell.alignment = _center()
                cell.font = Font(name="Arial", size=10, color="888888")
            elif ci == 2:
                cell.number_format = "DD/MM/YYYY"; cell.alignment = _center()
            elif ci in (4, 5):
                cell.number_format = "#,##0.00"; cell.alignment = _right()
            elif ci in (9, 10, 11):
                cell.alignment = _center()
            elif ci == 12:
                cell.number_format = "DD/MM/YYYY"; cell.alignment = _center()

            # STATUS colorido
            if ci == 11:
                status_val = str(v).upper()
                if status_val == "PAGO":
                    cell.font = Font(name="Arial", size=10, bold=True, color="4CAF7D")
                elif status_val == "PENDENTE":
                    cell.font = Font(name="Arial", size=10, bold=True, color="F05454")

    return next_id + len(df_new) - 1


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Processa extratos RAW_ e atualiza DB_DESPESAS.")
    parser.add_argument("--arquivo", default=str(DEFAULT_FILE), help="Caminho para o .xlsx")
    parser.add_argument("--banco",   default="todos",           help="Banco(s) a processar: C6,NUBANK,... ou 'todos'")
    args = parser.parse_args()

    filepath = Path(args.arquivo)
    if not filepath.exists():
        print(f"❌ Arquivo não encontrado: {filepath}")
        sys.exit(1)

    bancos_alvo = (
        BANCOS_DISPONIVEIS if args.banco.lower() == "todos"
        else [b.strip().upper() for b in args.banco.split(",")]
    )

    print(f"\n{'─'*60}")
    print(f"  Controle Financeiro — Processador de Extratos")
    print(f"  Arquivo : {filepath.name}")
    print(f"  Bancos  : {', '.join(bancos_alvo)}")
    print(f"  Data    : {date.today().strftime('%d/%m/%Y')}")
    print(f"{'─'*60}\n")

    xl      = pd.ExcelFile(filepath)
    regras  = pd.read_excel(xl, "REGRAS_CATEGORIAS", header=0)
    regras.columns = regras.columns.str.strip()

    # Lê DB_DESPESAS existente para deduplicação e ID máximo
    df_db = pd.read_excel(xl, "DB_DESPESAS", header=0)
    df_db.columns = df_db.columns.str.strip()
    df_db["Data Lançamento"] = pd.to_datetime(df_db["Data Lançamento"], errors="coerce")
    df_db["Saída(R$)"]       = pd.to_numeric(df_db["Saída(R$)"], errors="coerce").fillna(0)
    max_id = int(df_db["ID"].dropna().max()) if "ID" in df_db.columns and not df_db["ID"].dropna().empty else 0

    wb = openpyxl.load_workbook(filepath)
    total_inseridas = 0

    for banco in bancos_alvo:
        print(f"🏦 Processando {banco}...")
        try:
            df_raw  = ler_raw_sheet(xl, banco)
            if df_raw.empty:
                print(f"  ℹ️  RAW_{banco} está vazia — nenhuma linha processada.\n")
                continue

            norm_fn = NORMALIZADORES.get(banco)
            if not norm_fn:
                print(f"  ⚠️  Normalizador não encontrado para {banco}.\n")
                continue

            df_norm  = norm_fn(df_raw)
            df_final = aplicar_regras(df_norm, regras)
            df_final = deduplicate(df_db, df_final)

            if df_final.empty:
                print(f"  ✅ Nenhuma linha nova — tudo já estava em DB_DESPESAS.\n")
                continue

            print(f"  ➕ {len(df_final)} nova(s) linha(s) a inserir.")
            max_id = escrever_db_despesas(wb, df_final, max_id)

            # Atualiza df_db para próxima iteração de deduplicação
            df_db = pd.concat([df_db, df_final], ignore_index=True)
            total_inseridas += len(df_final)
            print(f"  ✅ {banco} concluído.\n")

        except Exception as e:
            print(f"  ❌ Erro ao processar {banco}: {e}\n")

    if total_inseridas > 0:
        wb.save(filepath)
        print(f"{'─'*60}")
        print(f"  💾 {total_inseridas} linha(s) inserida(s) em DB_DESPESAS.")
        print(f"  💾 Arquivo salvo: {filepath.name}")
    else:
        print(f"{'─'*60}")
        print(f"  ℹ️  Nenhuma linha nova inserida. Arquivo não alterado.")

    print(f"{'─'*60}\n")
    print("  ➡️  Próximos passos:")
    print("     1. Abra o dashboard: streamlit run dashboard.py")
    print("     2. Preencha DESC. BASE nas linhas novas (coluna amarela)")
    print("     3. Adicione novas regras em REGRAS_CATEGORIAS se necessário")
    print()


if __name__ == "__main__":
    main()
